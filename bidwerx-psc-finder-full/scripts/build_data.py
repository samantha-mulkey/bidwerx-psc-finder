#!/usr/bin/env python3
"""Normalize the official Acquisition.gov PSC workbook into active 4-character PSC records."""
import json
from datetime import date, datetime
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / 'data'

# Change this filename when Acquisition.gov publishes a newer workbook.
WORKBOOK = DATA / 'PSC_April_2025.xlsx'
OUTPUT = DATA / 'psc-normalized.json'
SOURCE_URL = 'https://www.acquisition.gov/sites/default/files/manual/PSC%20April%202025.xlsx'
SOURCE_DATE = '2025-04-08'


def text(value):
    if value is None:
        return ''
    return ' '.join(str(value).split())


def code_text(value):
    if value is None:
        return ''
    if isinstance(value, bool):
        return ''
    if isinstance(value, (int, float)):
        if float(value).is_integer():
            return str(int(value))
    return text(value).upper()


def json_date(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return text(value)


wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
rows = ws.iter_rows(values_only=True)
headers = [text(v) for v in next(rows)]
index = {name: i for i, name in enumerate(headers)}

records = []
for row in rows:
    code = code_text(row[index['PSC CODE']])
    end_date = row[index['END DATE']]

    # The finder is intentionally limited to active, assignable 4-character PSCs.
    if len(code) != 4 or end_date not in (None, ''):
        continue

    title = text(row[index['PRODUCT AND SERVICE CODE NAME']])
    if not title:
        continue

    records.append({
        'code': code,
        'title': title,
        'description': text(row[index['PRODUCT AND SERVICE CODE FULL NAME (DESCRIPTION)']]),
        'includes': text(row[index['PRODUCT AND SERVICE CODE INCLUDES']]),
        'excludes': text(row[index['PRODUCT AND SERVICE CODE EXCLUDES']]),
        'notes': text(row[index['PRODUCT AND SERVICE CODE NOTES']]),
        'parent': text(row[index['Parent PSC Code']]),
        'type': text(row[index['PSC Category: Service (S)/Product (P)']]),
        'level1_code': text(row[index['Level 1 Category Code']]),
        'level1_category': text(row[index['Level 1 Category']]),
        'level2_code': text(row[index['Level 2 Category Code']]),
        'level2_category': text(row[index['Level 2 Category']]),
    })

records.sort(key=lambda r: r['code'])

payload = {
    'source': SOURCE_URL,
    'source_date': SOURCE_DATE,
    'record_count': len(records),
    'records': records,
}

with OUTPUT.open('w', encoding='utf-8') as f:
    json.dump(payload, f, indent=2, ensure_ascii=False)

print(f'Wrote {len(records)} active PSC records to {OUTPUT}')
